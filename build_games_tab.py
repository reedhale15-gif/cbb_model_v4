import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
from openpyxl.formatting.rule import FormulaRule
from openpyxl.utils import get_column_letter

# =========================
# LOAD ENGINE DATA
# =========================

df = pd.read_csv("data/engine.csv")

# Remove games with no market spread
df = df[df["Spread"].notna()]

# Keep only Games tab columns
games_df = df[["Home", "Away", "Spread", "Total"]].copy()

# =========================
# CREATE WORKBOOK
# =========================

wb = Workbook()
ws = wb.active
ws.title = "Games"

headers = ["Home", "Away", "Spread", "Total"]
ws.append(headers)

# Bold header
for col in range(1, len(headers) + 1):
    ws.cell(row=1, column=col).font = Font(bold=True)

# Freeze header
ws.freeze_panes = "A2"

# =========================
# WRITE DATA
# =========================

for _, row in games_df.iterrows():
    ws.append(row.tolist())

# Bold Home column
for row in range(2, ws.max_row + 1):
    ws.cell(row=row, column=1).font = Font(bold=True)

# Alternating row shading
gray_fill = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")

for row in range(2, ws.max_row + 1):
    if row % 2 == 0:
        for col in range(1, 5):
            ws.cell(row=row, column=col).fill = gray_fill

# =========================
# CONDITIONAL FORMATTING
# =========================

orange_fill = PatternFill(start_color="FFE5CC", end_color="FFE5CC", fill_type="solid")
blue_fill = PatternFill(start_color="DDEBF7", end_color="DDEBF7", fill_type="solid")

rule_high = FormulaRule(formula=["$D2>155"], fill=orange_fill)
rule_low = FormulaRule(formula=["$D2<135"], fill=blue_fill)

ws.conditional_formatting.add(f"D2:D{ws.max_row}", rule_high)
ws.conditional_formatting.add(f"D2:D{ws.max_row}", rule_low)

# =========================
# COLUMN WIDTH
# =========================

for col in range(1, 5):
    ws.column_dimensions[get_column_letter(col)].width = 18

# =========================
# SAVE
# =========================

wb.save("data/games.xlsx")

print("Games tab exported to data/games.xlsx")
