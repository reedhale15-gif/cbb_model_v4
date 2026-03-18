from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter
import pandas as pd
from model_config import TOTAL_EDGE_MAX, TOTAL_EDGE_MIN, spread_bet_qualifies, spread_edge_band

EDGE_THRESHOLD_SPREAD, MAX_SPREAD_EDGE = spread_edge_band()
EDGE_THRESHOLD_TOTAL = TOTAL_EDGE_MIN

# ----------------------------
# Spread Confidence (Edge-Based)
# ----------------------------

def spread_confidence(edge):
    edge = abs(edge)
    if edge >= 12: return "A+"
    if edge >= 10: return "A"
    if edge >= 8: return "A-"
    if edge >= 6: return "B"
    if edge >= 5: return "C+"
    if edge >= 4: return "C"
    return ""

# ----------------------------
# Total Confidence (Keep Existing Ladder)
# ----------------------------

def total_confidence(edge):
    edge = float(edge)
    edge_abs = abs(edge)

    if edge > 0:
        if edge_abs < 10: return "A"
        if edge_abs <= 12: return "B"
        return ""

    if edge_abs >= 10: return "A"
    if edge_abs >= 6:
        if edge_abs < 8: return "B"
        return "C"

    return ""

df = pd.read_csv("data/engine.csv")

rows = []

for _, row in df.iterrows():

    home = row["Home"]
    away = row["Away"]
    market_spread = float(row["Spread"])
    model_spread = float(row["Model Spread"])
    spread_edge = float(row["Spread Edge"])
    market_total = float(row["Total"])
    model_total = float(row["Model Total"])
    total_edge = float(row["Total Edge"])

    # -------- Spread Bet Logic --------
    if spread_bet_qualifies(market_spread, model_spread, spread_edge):
        if model_spread < market_spread:
            spread_bet = f"{home} {market_spread:+.1f}"
        else:
            spread_bet = f"{away} {-market_spread:+.1f}"
        spread_conf = spread_confidence(spread_edge)
    else:
        spread_bet = "No Bet"
        spread_conf = ""

    # -------- Total Bet Logic --------
    if EDGE_THRESHOLD_TOTAL <= abs(total_edge) <= TOTAL_EDGE_MAX:
        if total_edge > 0:
            total_bet = f"Over {market_total:.1f}"
        else:
            total_bet = f"Under {market_total:.1f}"
        total_conf = total_confidence(total_edge)
    else:
        total_bet = "No Bet"
        total_conf = ""
        
    rows.append({
        "Home": home,
        "Away": away,
        "Spread": f"{market_spread:+.1f}",
        "Model Spread": f"{model_spread:+.1f}",
        "Spread Edge": f"{spread_edge:+.2f}",
        "Spread Bet": spread_bet,
        "Spread Confidence": spread_conf,
        "Total": f"{market_total:.1f}",
        "Model Total": f"{model_total:.1f}",
        "Total Edge": f"{total_edge:.2f}",
        "Total Bet": total_bet,
        "Total Confidence": total_conf
    })

bets_df = pd.DataFrame(rows)

wb = Workbook()
ws = wb.active
ws.title = "Bets"

headers = list(bets_df.columns)
ws.append(headers)

for col in range(1, len(headers)+1):
    ws.cell(row=1, column=col).font = Font(bold=True)

ws.freeze_panes = "A2"

for _, r in bets_df.iterrows():
    ws.append(r.tolist())

wb.save("data/bets.xlsx")

print("Bets tab exported to data/bets.xlsx")
